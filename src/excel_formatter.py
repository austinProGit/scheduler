# Max Lewis 09/20/22
# CPSC 4175 Project

import openpyxl
import shutil
from datetime import date
from driver_fs_functions import *


def excel_formatter(input_path, output_file_name, sched, container):  # first param is source file path
    num = 0
    input_file_name = get_source_relative_path(input_path, 'Path To Graduation Y.xlsx')
    for lst in sched:
        if len(lst) > 6:
            input_file_name = get_source_relative_path(input_path, 'Path To Graduation Z.xlsx')
            num = 4
            break

    shutil.copy2(input_file_name, output_file_name) # Copy file
    edit_excel = openpyxl.load_workbook(output_file_name)
    sheet = edit_excel.active
    format(sheet, sched, num, container)
    edit_excel.save(output_file_name) # Save the file so we can copy to specified directory
 
# ...............HELPER METHODS...............HELPER METHODS...............HELPER METHODS...............HELPER METHODS
    
def format(sheet, sched, num, container):
    stub_list = []
    easy_sched = []

    for semester in sched:
        semester_sched = []

        for schedulable in semester:
            semester_sched.append(str(schedulable.course_identifier))

            if schedulable.course_identifier.is_stub():
                stub_list.append(str(schedulable.course_identifier))

        easy_sched.append(semester_sched)

    current_seas = [1, 4] # We can switch to actual date reference by using current_season()
    for i in range(len(easy_sched)):
        if(i > 0): # Skip 1st iteration before we change semester to next
            current_seas = next_season(current_seas, num)

        for x in range(len(easy_sched[i])):
            course = easy_sched[i][x]
            data = ""
            hours = 3

            if course in stub_list:
                data = f"{course} - Elective (Fa Sp --)"
            else:
                name = container.get_name(course)
                avail = container.get_availability(course)
                data = f"{course} - {name} ({avail[0]} {avail[1]} {avail[2]})"
                hours = container.get_hours(course)

            co = current_seas[0]
            ro = current_seas[1] + x
            c = current_seas[0] + 1
            r = current_seas[1] + x
            sheet.cell(row=ro, column=co, value=data)
            sheet.cell(row=r, column=c, value=hours)


def current_season():
    start =[]
    month = date.today().month
    if month >= 1 and month <= 5: start = [5, 4] # [column, row] Summer table
    if month >= 6 and month <= 7: start = [1, 4] # [column, row] Fall table
    if month >= 8 and month <= 12: start = [3, 4] # [column, row] Spring table
    return start


def next_season(current_season, num):
    next_season = []
    if current_season[0] == 5:
        next_row = current_season[1] + 8 + num
        next_season = [1, next_row]
    if current_season[0] == 1:
        next_season = [3, current_season[1]]
    if current_season[0] == 3:
        next_season = [5, current_season[1]]
    return next_season

# ...............END OF HELPER METHODS...............END OF HELPER METHODS...............END OF HELPER METHODS

# ***************END OF EXCEL_FORMATTER MODULE***************END OF EXCEL_FORMATTER MODULE********************
#Internal testing below