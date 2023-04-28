#overwrites a path to graduations stubbed elective courses with electives obtained from cbr result

from openpyxl import Workbook, load_workbook

def write_to_file(selected_file, working_elective_list):
    electives_list_to_be_added =[]
    for course in working_elective_list:
         electives_list_to_be_added.append(course)
    write_count = 0
    wb = load_workbook(selected_file)
    ws = wb.active
    #iterate rows
    for i in range(1, 43):
        #iterate columns
        for j in range (1, 7):
            #get cell value 
            cell_obj = ws.cell(row=i, column=j)
            if (cell_obj.value == "Course XXXX - Elective (Fa Sp --)" and len(electives_list_to_be_added) > 0):
                    cell_obj.value = electives_list_to_be_added.pop()
                    write_count += 1
            elif len(electives_list_to_be_added) == 0:
                break
    wb.save(selected_file)
    return "Done - " + str(write_count) + " electives have been scheduled"
                
